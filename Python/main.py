from openpyxl import load_workbook
import datetime
class Info:
    def __init__(self):

        Name = input("Enter your FULL NAME: ")
        self.n = Name
        NIC = input("Enter your CNIC NUMBER: ")
        self.CNIC = NIC
        while len(self.CNIC) < 13 or len(self.CNIC) > 13:
            self.CNIC = input("Please Enter Valid CNIC NUMBER: ")
        M_NO = input("Enter your MOBILE NUMBER: ")
        if len(M_NO)==11:
            self.Mobile_NO = M_NO
        while len(M_NO) < 11 or len(M_NO) > 11:
            M_NO = input("Enter your VALID MOBILE NUMBER: ")
            if len(M_NO) == 11:
                self.Mobile_NO = M_NO

class City(Info):
    def select_city(self):
        self.wb=load_workbook("FINAL.xlsx")
        print("Cities in which our Cinemas are available:")
        c=self.wb['City-Cinema']
        self.city_list=[]
        for i in range(1,5):
            for j in range(1,2):
                a=(c.cell(row=i,column=j).value)
                print(a)
                self.city_list.append(a)
        self.city = input("Enter City: ")
        if self.city in self.city_list:
            print("City has been selected successfully!")
        else:
            print("SORRY! our Cinemas are not available in ",self.city)
        while self.city not in self.city_list:
            print("Our Cinemas are only available in:")
            for i in range(1, 5):
                for j in range(1, 2):
                    a = (c.cell(row=i, column=j).value)
                    print(a)
            self.city = input("Please enter the city from the above given options: ")
            if self.city in self.city_list:
                print("City has been selected successfully!")
            else:
                print("SORRY! our Cinemas are not available in ",self.city)


class Area(City):
    def select_area(self):
        cinema_name =self.wb['City-Cinema']
        if self.city=="KARACHI":
            K_area={"1":"K1","2":"K2","3":"K3","4":"K4"}
            print("These are the Areas of KARACHI in which our Cinemas are available!")
            self.area_list=[]
            show=[]
            for x,y in K_area.items():
                print(x+": "+y)
                self.area_list.append(x)
                show.append(y)
            self.area=str(input("Select your area 1,2,3 or 4: "))
            if self.area in self.area_list:
                print("Your area has been selected!")
                if self.area == "1":
                    self.show_area= show[0]
                    self.cin_name=cinema_name.cell(row=1, column=2).value
                    print("Name of Cinema",self.cin_name)
                elif self.area == "2":
                    self.show_area = show[1]
                    self.cin_name = cinema_name.cell(row=1, column=3).value
                    print("Name of Cinema",self.cin_name)
                elif self.area == "3":
                    self.show_area = show[2]
                    self.cin_name = cinema_name.cell(row=1, column=4).value
                    print("Name of Cinema",self.cin_name)
                elif self.area == "4":
                    self.show_area = show[3]
                    self.cin_name = cinema_name.cell(row=1, column=5).value
                    print("Name of Cinema",self.cin_name)
            else:
                print("Sorry! our service is not availble here")

            while self.area not in self.area_list:
                for x, y in K_area.items():
                    print(x + ": " + y)
                self.area = str(input("Please select your area from above given options 1,2,3 or 4: "))
                if self.area in self.area_list:
                    print("Your area has been selected!")
                    if self.area == "1":
                        self.show_area = show[0]
                        self.cin_name = cinema_name.cell(row=1, column=2).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "2":
                        self.show_area = show[1]
                        self.cin_name = cinema_name.cell(row=1, column=3).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "3":
                        self.show_area = show[2]
                        self.cin_name = cinema_name.cell(row=1, column=4).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "4":
                        self.show_area = show[3]
                        self.cin_name = cinema_name.cell(row=1, column=5).value
                        print("Name of Cinema", self.cin_name)
                else:
                    print("Sorry! our service is not availble here")


        elif self.city=="LAHORE":
            L_area={"1":"L1","2":"L2","3":"L3","4":"L4"}
            print("These are the Areas of LAHORE in which our Cinemas are available!")
            self.area_list=[]
            show = []
            for x,y in L_area.items():
                print(x+": "+y)
                self.area_list.append(x)
                show.append(y)
            self.area=str(input("Select your area 1,2,3 or 4: "))
            if self.area in self.area_list:
                print("Your area has been selected!")
                if self.area == "1":
                    self.show_area = show[0]
                    self.cin_name=cinema_name.cell(row=2, column=2).value
                    print("Name of Cinema",self.cin_name)
                elif self.area == "2":
                    self.show_area = show[1]
                    self.cin_name = cinema_name.cell(row=2, column=3).value
                    print("Name of Cinema",self.cin_name)
                elif self.area == "3":
                    self.show_area = show[2]
                    self.cin_name = cinema_name.cell(row=2, column=4).value
                    print("Name of Cinema",self.cin_name)
                elif self.area == "4":
                    self.show_area = show[3]
                    self.cin_name = cinema_name.cell(row=2, column=5).value
                    print("Name of Cinema",self.cin_name)
            else:
                print("Sorry! our service is not availble here")
            while self.area not in self.area_list:
                for x, y in L_area.items():
                    print(x + ": " + y)
                self.area = str(input("Please select your area from above given options 1,2,3 or 4: "))
                if self.area in self.area_list:
                    print("Your area has been selected")
                    if self.area == "1":
                        self.show_area = show[0]
                        self.cin_name = cinema_name.cell(row=2, column=2).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "2":
                        self.show_area = show[1]
                        self.cin_name = cinema_name.cell(row=2, column=3).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "3":
                        self.show_area = show[2]
                        self.cin_name = cinema_name.cell(row=2, column=4).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "4":
                        self.show_area = show[3]
                        self.cin_name = cinema_name.cell(row=2, column=5).value
                        print("Name of Cinema", self.cin_name)
                else:
                    print("Sorry! our service is not availble here")


        elif self.city=="ISLAMABAD":
                I_area={"1":"I1","2":"I2","3":"I3","4":"I4"}
                print("These are the Areas of ISLAMABAD in which our Cinemas are available!")
                self.area_list=[]
                show=[]
                for x,y in I_area.items():
                    print(x+": "+y)
                    self.area_list.append(x)
                    show.append(y)
                self.area=str(input("Select your area 1,2,3 or 4: "))
                if self.area in self.area_list:
                    print("Your area has been selected!")
                    if self.area == "1":
                        self.show_area = show[0]
                        self.cin_name=cinema_name.cell(row=3, column=2).value
                        print("Name of Cinema",self.cin_name)
                    elif self.area == "2":
                        self.show_area = show[1]
                        self.cin_name = cinema_name.cell(row=3, column=3).value
                        print("Name of Cinema",self.cin_name)
                    elif self.area == "3":
                        self.show_area = show[2]
                        self.cin_name = cinema_name.cell(row=3, column=4).value
                        print("Name of Cinema",self.cin_name)
                    elif self.area == "4":
                        self.show_area = show[3]
                        self.cin_name = cinema_name.cell(row=3, column=5).value
                        print("Name of Cinema",self.cin_name)
                else:
                    print("Sorry! our service is not availble here")
                while self.area not in self.area_list:
                    for x, y in I_area.items():
                        print(x + ": " + y)
                    self.area = str(input("Please select your area from above given options 1,2,3 or 4: "))
                    if self.area in self.area_list:
                        print("Your area has been selected")
                        if self.area == "1":
                            self.show_area = show[0]
                            self.cin_name = cinema_name.cell(row=3, column=2).value
                            print("Name of Cinema", self.cin_name)
                        elif self.area == "2":
                            self.show_area = show[1]
                            self.cin_name = cinema_name.cell(row=3, column=3).value
                            print("Name of Cinema", self.cin_name)
                        elif self.area == "3":
                            self.show_area = show[2]
                            self.cin_name = cinema_name.cell(row=3, column=4).value
                            print("Name of Cinema", self.cin_name)
                        elif self.area == "4":
                            self.show_area = show[3]
                            self.cin_name = cinema_name.cell(row=3, column=5).value
                            print("Name of Cinema", self.cin_name)
                    else:
                        print("Sorry! our service is not availble here")


        elif self.city=="QUETTA":
                Q_area={"1":"Q1","2":"Q2","3":"Q3","4":"Q4"}
                print("These are the Areas of QUETTA in which our Cinemas are available!")
                self.area_list=[]
                show = []
                for x,y in Q_area.items():
                    print(x+": "+y)
                    self.area_list.append(x)
                    show.append(y)
                self.area=str(input("Select your area 1,2,3 or 4: "))
                if self.area in self.area_list:
                    print("Your area has been selected!")
                    if self.area == "1":
                        self.show_area = show[0]
                        self.cin_name=cinema_name.cell(row=4, column=2).value
                        print("Name of Cinema",self.cin_name)
                    elif self.area == "2":
                        self.show_area = show[1]
                        self.cin_name=cinema_name.cell(row=4, column=3).value
                        print("Name of Cinema",self.cin_name)
                    elif self.area == "3":
                        self.show_area = show[2]
                        self.cin_name=cinema_name.cell(row=4, column=4).value
                        print("Name of Cinema",self.cin_name)
                    elif self.area == "4":
                        self.show_area = show[3]
                        self.cin_name=cinema_name.cell(row=4, column=5).value
                        print("Name of Cinema",self.cin_name)
                else:
                    print("Sorry! our service is not availble here")
                while self.area not in self.area_list:
                    for x, y in Q_area.items():
                        print(x + ": " + y)
                    self.area = str(input("Please select your area from above given options 1,2,3 or 4: "))
                    if self.area == "1":
                        self.show_area = show[0]
                        self.cin_name=cinema_name.cell(row=4, column=2).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "2":
                        self.show_area = show[1]
                        self.cin_name=cinema_name.cell(row=4, column=3).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "3":
                        self.cin_name=self.show_area = show[2]
                        cinema_name.cell(row=4, column=4).value
                        print("Name of Cinema", self.cin_name)
                    elif self.area == "4":
                        self.show_area = show[3]
                        self.cin_name=cinema_name.cell(row=4, column=5).value
                        print("Name of Cinema", self.cin_name)
                    else:
                        print("Sorry! our service is not availble here")


class Movies(Area):
    def select_movie(self):
        print("MOVIES which are avalable:")
        c = self.wb['Movies']
        self.movies_list = []
        keys = ["1", "2", "3"]
        self.movies_dict = {}
        self.key_list = []
        for i in range(1, 4):
            for j in range(1, 2):
                a = (c.cell(row=i, column=j).value)
                self.movies_list.append(a)
        for k in range(len(keys)):
            self.movies_dict[keys[k]] = self.movies_list[k]
        for x, y in self.movies_dict.items():
            print(x + ": " + y)
            self.key_list.append(x)
        self.movie = input("Select Movie 1,2 or 3: ")
        if self.movie in self.key_list:
            print("MOVIE has been selected successfully!")
        else:
            print("This movie is not available")
        while self.movie not in self.key_list:
            for x, y in self.movies_dict.items():
                print(x + ": " + y)
            self.movie = input("Please select movie from the above given options 1,2 or 3: ")
            if self.movie in self.key_list:
                print("MOVIE has been selected successfully!")
            else:
                print("This movie is not available")


class Screen_and_Time(Movies):
    def select_screen_time(self):
        s1 = self.wb['Timing']
        if (self.movie) == (self.key_list[0]):
            self.show_movie= self.movies_dict["1"]
            print(self.show_movie+" movie timing on different screen:")
            for j in range(1, 5):
                print(s1.cell(row=1, column=j).value)
            self.screen = int(input("Select screen 1,2 or 3: "))
            if self.screen == 1:
                self.show_time=s1['B1'].value
                print("SCREEN selected! Your timing on SCREEN 1 is: ", self.show_time)
            elif self.screen == 2:
                self.show_time = s1['C1'].value
                print("SCREEN selected! Your timing on SCREEN 2 is: ", self.show_time)
            elif self.screen == 3:
                self.show_time = s1['D1'].value
                print("SCREEN selected! Your timing on SCREEN 3 is: ", self.show_time)
            else:
                print("This screen is not available")
            s_list = [1, 2, 3]
            while (self.screen) not in s_list:
                self.screen = int(input("Please select screen 1,2 or 3 : "))
                if self.screen == 1:
                    self.show_time = s1['B1'].value
                    print("SCREEN selected! Your timing on SCREEN 1 is: ", self.show_time)
                elif self.screen == 2:
                    self.show_time = s1['C1'].value
                    print("SCREEN selected! Your timing on SCREEN 2 is: ", self.show_time)
                elif self.screen == 3:
                    self.show_time = s1['D1'].value
                    print("SCREEN selected! Your timing on SCREEN 3 is: ", self.show_time)
                else:
                    print("This screen is not available")

        elif (self.movie) == (self.key_list[1]):
            self.show_movie = self.movies_dict["2"]
            print(self.show_movie + " movie timing on different screen:")
            for k in range(1, 5):
                print(s1.cell(row=2, column=k).value)
            self.screen = int(input("Select screen: "))
            if self.screen == 1:
                self.show_time = s1['D2'].value
                print("SCREEN selected! Your timing on SCREEN 1 is: ", self.show_time)
            elif self.screen == 2:
                self.show_time = s1['B2'].value
                print("SCREEN selected! Your timing on SCREEN 2 is: ", self.show_time)
            elif self.screen == 3:
                self.show_time = s1['C2'].value
                print("SCREEN selected! Your timing on SCREEN 3 is: ", self.show_time)
            else:
                print("This screen is not available")
            s_list = [1, 2, 3]
            while (self.screen) not in s_list:
                self.screen = int(input("Please select screen 1,2 or 3: "))
                if self.screen == 1:
                    self.show_time = s1['D2'].value
                    print("SCREEN selected! Your timing on SCREEN 1 is: ", self.show_time)
                elif self.screen == 2:
                    self.show_time = s1['B2'].value
                    print("SCREEN selected! Your timing on SCREEN 2 is: ", self.show_time)
                elif self.screen == 3:
                    self.show_time = s1['C2'].value
                    print("SCREEN selected! Your timing on SCREEN 3 is: ", self.show_time)
                else:
                    print("This screen is not available")
        elif (self.movie) == (self.key_list[2]):
            self.show_movie = self.movies_dict["3"]
            print(self.show_movie + " movie timing on different screen:")
            for k in range(1, 5):
                print(s1.cell(row=3, column=k).value)
            self.screen = int(input("Select screen: "))
            if self.screen == 1:
                self.show_time = s1['C3'].value
                print("SCREEN selected! Your timing on SCREEN 1 is: ", self.show_time)
            elif self.screen == 2:
                self.show_time = s1['D3'].value
                print("SCREEN selected! Your timing on SCREEN 2 is: ", self.show_time)
            elif self.screen == 3:
                self.show_time = s1['B3'].value
                print("SCREEN selected! Your timing on SCREEN 3 is: ", self.show_time)
            else:
                print("This screen is not available")
            s_list = [1, 2, 3]
            while (self.screen) not in s_list:
                self.screen = int(input("Please select screen 1,2 or 3: "))
                if self.screen == 1:
                    self.show_time = s1['C3'].value
                    print("SCREEN selected! Your timing on SCREEN 1 is: ", self.show_time)
                elif self.screen == 2:
                    self.show_time = s1['D3'].value
                    print("SCREEN selected! Your timing on SCREEN 2 is: ", self.show_time)
                elif self.screen == 3:
                    self.show_time = s1['B3'].value
                    print("SCREEN selected! Your timing on SCREEN 3 is: ", self.show_time)
                else:
                    print("This screen is not available")


class Ticket_and_Seat(Screen_and_Time):
    def ticket(self):
        def cat_premium(self):
            for i in range(self.ticket):
                self.seat = int(input("Choose seat between 1-50: "))
                if self.seat >= 1 and self.seat <= 50:
                    if self.movie == "1" and self.screen == 1:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=3).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=3).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=3, value="BOOKED")
                                    e1.cell(row=self.seat, column=4, value=self.n)
                                    e1.cell(row=self.seat, column=5, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")

                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=3, value="BOOKED")
                            e1.cell(row=self.seat, column=4, value=self.n)
                            e1.cell(row=self.seat, column=5, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "1" and self.screen == 2:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=8).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=8).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=8, value="BOOKED")
                                    e1.cell(row=self.seat, column=9, value=self.n)
                                    e1.cell(row=self.seat, column=10, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=8, value="BOOKED")
                            e1.cell(row=self.seat, column=9, value=self.n)
                            e1.cell(row=self.seat, column=10, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "1" and self.screen == 3:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=13).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=13).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=13, value="BOOKED")
                                    e1.cell(row=self.seat, column=14, value=self.n)
                                    e1.cell(row=self.seat, column=15, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=13, value="BOOKED")
                            e1.cell(row=self.seat, column=14, value=self.n)
                            e1.cell(row=self.seat, column=15, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "2" and self.screen == 1:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=18).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=18).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=18, value="BOOKED")
                                    e1.cell(row=self.seat, column=19, value=self.n)
                                    e1.cell(row=self.seat, column=20, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=18, value="BOOKED")
                            e1.cell(row=self.seat, column=19, value=self.n)
                            e1.cell(row=self.seat, column=20, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "2" and self.screen == 2:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=23).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=23).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=23, value="BOOKED")
                                    e1.cell(row=self.seat, column=24, value=self.n)
                                    e1.cell(row=self.seat, column=25, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=23, value="BOOKED")
                            e1.cell(row=self.seat, column=24, value=self.n)
                            e1.cell(row=self.seat, column=25, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "2" and self.screen == 3:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=28).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=28).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=28, value="BOOKED")
                                    e1.cell(row=self.seat, column=29, value=self.n)
                                    e1.cell(row=self.seat, column=30, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=28, value="BOOKED")
                            e1.cell(row=self.seat, column=29, value=self.n)
                            e1.cell(row=self.seat, column=30, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "3" and self.screen == 1:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=33).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=33).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=33, value="BOOKED")
                                    e1.cell(row=self.seat, column=34, value=self.n)
                                    e1.cell(row=self.seat, column=35, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=33, value="BOOKED")
                            e1.cell(row=self.seat, column=34, value=self.n)
                            e1.cell(row=self.seat, column=35, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")

                    elif self.movie == "3" and self.screen == 2:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=38).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=38).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=38, value="BOOKED")
                                    e1.cell(row=self.seat, column=39, value=self.n)
                                    e1.cell(row=self.seat, column=40, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=38, value="BOOKED")
                            e1.cell(row=self.seat, column=39, value=self.n)
                            e1.cell(row=self.seat, column=40, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")

                    elif self.movie == "3" and self.screen == 3:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=43).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=43).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=43, value="BOOKED")
                                    e1.cell(row=self.seat, column=44, value=self.n)
                                    e1.cell(row=self.seat, column=45, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=43, value="BOOKED")
                            e1.cell(row=self.seat, column=44, value=self.n)
                            e1.cell(row=self.seat, column=45, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")

                while self.seat < 1 or self.seat > 50:
                    self.seat = int(input("Please Choose seat between 1-50: "))
                    if self.seat >= 1 and self.seat <= 50:
                        if self.movie == "1" and self.screen == 1:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=3).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=3).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=3, value="BOOKED")
                                        e1.cell(row=self.seat, column=4, value=self.n)
                                        e1.cell(row=self.seat, column=5, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")

                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=3, value="BOOKED")
                                e1.cell(row=self.seat, column=4, value=self.n)
                                e1.cell(row=self.seat, column=5, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "1" and self.screen == 2:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=8).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=8).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=8, value="BOOKED")
                                        e1.cell(row=self.seat, column=9, value=self.n)
                                        e1.cell(row=self.seat, column=10, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=8, value="BOOKED")
                                e1.cell(row=self.seat, column=9, value=self.n)
                                e1.cell(row=self.seat, column=10, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "1" and self.screen == 3:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=13).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=13).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=13, value="BOOKED")
                                        e1.cell(row=self.seat, column=14, value=self.n)
                                        e1.cell(row=self.seat, column=15, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=13, value="BOOKED")
                                e1.cell(row=self.seat, column=14, value=self.n)
                                e1.cell(row=self.seat, column=15, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "2" and self.screen == 1:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=18).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=18).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=18, value="BOOKED")
                                        e1.cell(row=self.seat, column=19, value=self.n)
                                        e1.cell(row=self.seat, column=20, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=18, value="BOOKED")
                                e1.cell(row=self.seat, column=19, value=self.n)
                                e1.cell(row=self.seat, column=20, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "2" and self.screen == 2:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=23).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=23).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=23, value="BOOKED")
                                        e1.cell(row=self.seat, column=24, value=self.n)
                                        e1.cell(row=self.seat, column=25, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=23, value="BOOKED")
                                e1.cell(row=self.seat, column=24, value=self.n)
                                e1.cell(row=self.seat, column=25, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "2" and self.screen == 3:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=28).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=28).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=28, value="BOOKED")
                                        e1.cell(row=self.seat, column=29, value=self.n)
                                        e1.cell(row=self.seat, column=30, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=28, value="BOOKED")
                                e1.cell(row=self.seat, column=29, value=self.n)
                                e1.cell(row=self.seat, column=30, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "3" and self.screen == 1:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=33).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=33).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=33, value="BOOKED")
                                        e1.cell(row=self.seat, column=34, value=self.n)
                                        e1.cell(row=self.seat, column=35, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=33, value="BOOKED")
                                e1.cell(row=self.seat, column=34, value=self.n)
                                e1.cell(row=self.seat, column=35, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")

                        elif self.movie == "3" and self.screen == 2:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=38).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=38).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=38, value="BOOKED")
                                        e1.cell(row=self.seat, column=39, value=self.n)
                                        e1.cell(row=self.seat, column=40, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=38, value="BOOKED")
                                e1.cell(row=self.seat, column=39, value=self.n)
                                e1.cell(row=self.seat, column=40, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")

                        elif self.movie == "3" and self.screen == 3:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=43).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=43).value
                                    if a != "BOOKED":
                                        print(self.seat, " seat selected!")
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=44, value=self.n)
                                        e1.cell(row=self.seat, column=45, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=43, value="BOOKED")
                                e1.cell(row=self.seat, column=44, value=self.n)
                                e1.cell(row=self.seat, column=45, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")

        def premium_payment_receipt(self):
            a = self.ticket * self.value[0]
            print("You have to pay ", a, "Rs")
            print("------Payment Method-------")
            print("We just have two payment methods")
            print("1 : Credit and Debit card")
            print("2 : Visa and Master card")
            A = input("Enter your METHOD of payment 1/2: ")
            if A == "1":
                B = input("Enter your Credit/Debit card number: ")
                print("Your seat reservation is confirmed")
                print("\t", "ONLINE CINEMA TICKET BOOKING IN PAKISTAN")
                print("\t", "ONLINE CINEMA TICKET BOOKING IN PAKISTAN")
                print("----------RECEIPT----------")
                print("NAME: " + self.n + "   CNIC: " + self.CNIC + "      MOBILE NUMBER: " + self.Mobile_NO)
                print("CITY: " + self.city + "     AREA: " + self.show_area + "     CINEMA NAME: " + self.cin_name)
                print("MOVIE: " + self.show_movie + "   SCREEN:", self.screen, "   TIME: ", self.show_time)
                print("SEAT CATEGORY: " + self.category + "  TOTAL SEATS: ", self.ticket)
                print("AMOUNT PAID: ", a, "   DATE:", datetime.date.today())
                print("\t", "\t", "www.ctbs.com.pk")
            elif A == "2":
                B = input("Enter your Visa/Master card number: ")
                print("your seat reservation is confirmed")
                print("\t", "ONLINE CINEMA TICKET BOOKING IN PAKISTAN")
                print("----------RECEIPT----------")
                print("NAME: " + self.n + "   CNIC: " + self.CNIC + "      MOBILE NUMBER: " + self.Mobile_NO)
                print("CITY: " + self.city + "     AREA: " + self.show_area + "     CINEMA NAME: " + self.cin_name)
                print("MOVIE: " + self.show_movie + "   SCREEN:", self.screen, "   TIME: ", self.show_time)
                print("SEAT CATEGORY: " + self.category + "  TOTAL SEATS: ", self.ticket)
                print("AMOUNT PAID: ", a, "   DATE:", datetime.date.today())
                print("\t", "\t", "www.ctbs.com.pk")

        def cat_normal(self):
            for i in range(self.ticket):
                self.seat = int(input("Choose seat between 51-100: "))
                if self.seat >= 51 and self.seat <= 100:
                    if self.movie == "1" and self.screen == 1:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=3).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=3).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=3, value="BOOKED")
                                    e1.cell(row=self.seat, column=4, value=self.n)
                                    e1.cell(row=self.seat, column=5, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")

                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=3, value="BOOKED")
                            e1.cell(row=self.seat, column=4, value=self.n)
                            e1.cell(row=self.seat, column=5, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "1" and self.screen == 2:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=8).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=8).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=8, value="BOOKED")
                                    e1.cell(row=self.seat, column=9, value=self.n)
                                    e1.cell(row=self.seat, column=10, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=8, value="BOOKED")
                            e1.cell(row=self.seat, column=9, value=self.n)
                            e1.cell(row=self.seat, column=10, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "1" and self.screen == 3:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=13).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=13).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=13, value="BOOKED")
                                    e1.cell(row=self.seat, column=14, value=self.n)
                                    e1.cell(row=self.seat, column=15, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=13, value="BOOKED")
                            e1.cell(row=self.seat, column=14, value=self.n)
                            e1.cell(row=self.seat, column=15, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "2" and self.screen == 1:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=18).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=18).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=18, value="BOOKED")
                                    e1.cell(row=self.seat, column=19, value=self.n)
                                    e1.cell(row=self.seat, column=20, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=18, value="BOOKED")
                            e1.cell(row=self.seat, column=19, value=self.n)
                            e1.cell(row=self.seat, column=20, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "2" and self.screen == 2:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=23).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=23).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=23, value="BOOKED")
                                    e1.cell(row=self.seat, column=24, value=self.n)
                                    e1.cell(row=self.seat, column=25, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=23, value="BOOKED")
                            e1.cell(row=self.seat, column=24, value=self.n)
                            e1.cell(row=self.seat, column=25, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "2" and self.screen == 3:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=28).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=28).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=28, value="BOOKED")
                                    e1.cell(row=self.seat, column=29, value=self.n)
                                    e1.cell(row=self.seat, column=30, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=28, value="BOOKED")
                            e1.cell(row=self.seat, column=29, value=self.n)
                            e1.cell(row=self.seat, column=30, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                    elif self.movie == "3" and self.screen == 1:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=33).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=33).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=33, value="BOOKED")
                                    e1.cell(row=self.seat, column=34, value=self.n)
                                    e1.cell(row=self.seat, column=35, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=33, value="BOOKED")
                            e1.cell(row=self.seat, column=34, value=self.n)
                            e1.cell(row=self.seat, column=35, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")

                    elif self.movie == "3" and self.screen == 2:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=38).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=38).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=38, value="BOOKED")
                                    e1.cell(row=self.seat, column=39, value=self.n)
                                    e1.cell(row=self.seat, column=40, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=38, value="BOOKED")
                            e1.cell(row=self.seat, column=39, value=self.n)
                            e1.cell(row=self.seat, column=40, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")

                    elif self.movie == "3" and self.screen == 3:
                        e1 = self.wb['SEAT']
                        a = e1.cell(row=self.seat, column=43).value
                        if a == "BOOKED":
                            print("This seat is already booked. Please choose a different one!")
                            while a == "BOOKED":
                                self.seat = int(input("Already BOOKED!Choose a different one: "))
                                a = e1.cell(row=self.seat, column=43).value
                                if a != "BOOKED":
                                    print("SEAT NUMBER: ",self.seat," SELECTED!")
                                    e1.cell(row=self.seat, column=43, value="BOOKED")
                                    e1.cell(row=self.seat, column=44, value=self.n)
                                    e1.cell(row=self.seat, column=45, value=self.CNIC)
                                    self.wb.save("FINAL.xlsx")
                        else:
                            print("SEAT NUMBER: ",self.seat," SELECTED!")
                            e1.cell(row=self.seat, column=43, value="BOOKED")
                            e1.cell(row=self.seat, column=44, value=self.n)
                            e1.cell(row=self.seat, column=45, value=self.CNIC)
                            self.wb.save("FINAL.xlsx")
                while self.seat < 50 or self.seat > 100:
                    self.seat = int(input("Please Choose seat between 51-100: "))
                    if self.seat >= 51 and self.seat <= 100:
                        if self.movie == "1" and self.screen == 1:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=3).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=3).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=3, value="BOOKED")
                                        e1.cell(row=self.seat, column=4, value=self.n)
                                        e1.cell(row=self.seat, column=5, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")

                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=3, value="BOOKED")
                                e1.cell(row=self.seat, column=4, value=self.n)
                                e1.cell(row=self.seat, column=5, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "1" and self.screen == 2:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=8).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=8).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=8, value="BOOKED")
                                        e1.cell(row=self.seat, column=9, value=self.n)
                                        e1.cell(row=self.seat, column=10, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=8, value="BOOKED")
                                e1.cell(row=self.seat, column=9, value=self.n)
                                e1.cell(row=self.seat, column=10, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "1" and self.screen == 3:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=13).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=13).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=13, value="BOOKED")
                                        e1.cell(row=self.seat, column=14, value=self.n)
                                        e1.cell(row=self.seat, column=15, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=13, value="BOOKED")
                                e1.cell(row=self.seat, column=14, value=self.n)
                                e1.cell(row=self.seat, column=15, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "2" and self.screen == 1:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=18).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=18).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=18, value="BOOKED")
                                        e1.cell(row=self.seat, column=19, value=self.n)
                                        e1.cell(row=self.seat, column=20, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=18, value="BOOKED")
                                e1.cell(row=self.seat, column=19, value=self.n)
                                e1.cell(row=self.seat, column=20, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "2" and self.screen == 2:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=23).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=23).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=23, value="BOOKED")
                                        e1.cell(row=self.seat, column=24, value=self.n)
                                        e1.cell(row=self.seat, column=25, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=23, value="BOOKED")
                                e1.cell(row=self.seat, column=24, value=self.n)
                                e1.cell(row=self.seat, column=25, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "2" and self.screen == 3:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=28).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=28).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=28, value="BOOKED")
                                        e1.cell(row=self.seat, column=29, value=self.n)
                                        e1.cell(row=self.seat, column=30, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=28, value="BOOKED")
                                e1.cell(row=self.seat, column=29, value=self.n)
                                e1.cell(row=self.seat, column=30, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")
                        elif self.movie == "3" and self.screen == 1:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=33).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=33).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=33, value="BOOKED")
                                        e1.cell(row=self.seat, column=34, value=self.n)
                                        e1.cell(row=self.seat, column=35, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=33, value="BOOKED")
                                e1.cell(row=self.seat, column=34, value=self.n)
                                e1.cell(row=self.seat, column=35, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")

                        elif self.movie == "3" and self.screen == 2:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=38).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=38).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=38, value="BOOKED")
                                        e1.cell(row=self.seat, column=39, value=self.n)
                                        e1.cell(row=self.seat, column=40, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=38, value="BOOKED")
                                e1.cell(row=self.seat, column=39, value=self.n)
                                e1.cell(row=self.seat, column=40, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")

                        elif self.movie == "3" and self.screen == 3:
                            e1 = self.wb['SEAT']
                            a = e1.cell(row=self.seat, column=43).value
                            if a == "BOOKED":
                                print("This seat is already booked. Please choose a different one!")
                                while a == "BOOKED":
                                    self.seat = int(input("Already BOOKED!Choose a different one: "))
                                    a = e1.cell(row=self.seat, column=43).value
                                    if a != "BOOKED":
                                        print("SEAT NUMBER: ",self.seat," SELECTED!")
                                        e1.cell(row=self.seat, column=43, value="BOOKED")
                                        e1.cell(row=self.seat, column=44, value=self.n)
                                        e1.cell(row=self.seat, column=45, value=self.CNIC)
                                        self.wb.save("FINAL.xlsx")
                            else:
                                print("SEAT NUMBER: ",self.seat," SELECTED!")
                                e1.cell(row=self.seat, column=43, value="BOOKED")
                                e1.cell(row=self.seat, column=44, value=self.n)
                                e1.cell(row=self.seat, column=45, value=self.CNIC)
                                self.wb.save("FINAL.xlsx")

        def normal_payment_receipt(self):
            a = self.ticket * self.value[1]
            print("You have to pay ", a, "Rs")
            print("------Payment Method-------")
            print("We just have two payment methods")
            print("1 : Credit and Debit card")
            print("2 : Visa and Master card")
            A = input("Enter your METHOD of payment 1/2: ")
            if A == "1":
                B = input("Enter your Credit/Debit card number: ")
                print("Your seat reservation is confirmed")
                print("\t", "ONLINE CINEMA TICKET BOOKING IN PAKISTAN")
                print("----------RECEIPT----------")
                print("NAME: " + self.n + "   CNIC: " + self.CNIC + "      MOBILE NUMBER: " + self.Mobile_NO)
                print("CITY: " + self.city + "     AREA: " + self.show_area + "     CINEMA NAME: " + self.cin_name)
                print("MOVIE: " + self.show_movie + "   SCREEN:", self.screen, "   TIME: ", self.show_time)
                print("SEAT CATEGORY: " + self.category + "  TOTAL SEATS: ", self.ticket)
                print("AMOUNT PAID: ", a, "   DATE:", datetime.date.today())
                print("\t", "\t", "www.ctbs.com.pk")
            elif A == "2":
                B = input("Enter your Visa/Master card number: ")
                print("your seat reservation is confirmed")
                print("\t", "ONLINE CINEMA TICKET BOOKING IN PAKISTAN")
                print("----------RECEIPT----------")
                print("NAME: " + self.n + "   CNIC: " + self.CNIC + "      MOBILE NUMBER: " + self.Mobile_NO)
                print("CITY: " + self.city + "     AREA: " + self.show_area + "     CINEMA NAME: " + self.cin_name)
                print("MOVIE: " + self.show_movie + "   SCREEN:", self.screen, "   TIME: ", self.show_time)
                print("SEAT CATEGORY: " + self.category + "  TOTAL SEATS: ", self.ticket)
                print("AMOUNT PAID: ", a, "   DATE:", datetime.date.today())
                print("\t", "\t", "www.ctbs.com.pk")
        key = ["PREMIUM", "NORMAL"]
        self.value = [1100, 800]
        key2 = [1, 2]
        self.ticket_dit = {}
        dict2 = {}
        print("Seat Arrangements are as follow:")
        print("PREMIUM= 1-50")
        print("NORMAL= 51-100")
        print("Price of each seat category wise:")
        for i in range(len(key)):
            self.ticket_dit[key[i]] = self.value[i]
        for j in range(len(key2)):
            dict2[key2[j]] = key[j]
        for x, y in self.ticket_dit.items():
            print(x, ": ", y, "Rs")
        print("BUY TICKETS!!")
        self.ticket = int(input("HOW MANY TICKETS YOU WANNA BUY? "))
        print(dict2)
        if self.ticket >20:
            print("CONTACT: 03131321173")
        elif self.ticket<20:
            self.category = input("Enter Category PREMIUM/NORMAL: ")
            if self.category == "PREMIUM":
                cat_premium(self)
                premium_payment_receipt(self)
            elif self.category == "NORMAL":

                cat_normal(self)
                normal_payment_receipt(self)
            else:
                print("*You enter wrong Category* *CHOOSE PREMIUM / NORMAL*")
        while self.category not in key:
            self.category=input("Enter Category PREMIUM/NORMAL: ")
            if self.category == "PREMIUM":
                cat_premium(self)
                premium_payment_receipt(self)
            elif self.category == "NORMAL":

                cat_normal(self)
                normal_payment_receipt(self)
            else:
                print("*You enter wrong Category* *CHOOSE PREMIUM / NORMAL*")





object=Ticket_and_Seat()
object.select_city()
object.select_area()
object.select_movie()
object.select_screen_time()
object.ticket()

