#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import openpyxl
def flight():
    airline=openpyxl.load_workbook("G:\\SIMPLE_AIRLINE_BOOKING_SYSTEM_IN_PYTHON_WITH_SOURCE_CODE\\SimpleAirlinesBookingSystem_PYTHON\\Airline-Booking-System\\Book1.xlsx")
    opt_1= 'Return'
    opt_2= 'Oneway'
    print(opt_1)
    print(opt_2)
    opt = input("Select opt_1 or opt_2:")
    if opt == opt_1:
        fro = input("From which state:")
        to = input("To which state:")
        depart_date = input("Enter Departure Date:")
        return_date = input("Enter Return Date:")
        cabin = input("Enter Cabin:")
        sh2=airline['FLIGHT']
        dt1 = sh2['A2'].value = fro
        dt2 = sh2['B2'].value = to
        dt3 = sh2['C2'].value = depart_date
        dt4 = sh2['D2'].value = return_date
        dt6 = sh2['F2'].value = cabin
        sh1=airline['Return']
        row=sh1.max_row
        column=sh1.max_column
        for i in range(1,row):
            for j in range(1,column):
                b=sh1.cell(i,j).value
                if b == 'Available':
          print(sh1.cell(i,1).value,sh1.cell(1,j).value,sh1.cell(i,j).value)
        Depart_Date2=input("Now Enter Your Departure Date:")
        Return_Date2=input("Now Enter Return Date:")
        payment=int(50000)
        Fst_Name =input("Enter First Name:")
        last_Name = input("Enter Last Name:")
        Sur_Name = input("Enter your SurName:")
        D_O_B = input("Enter Your Date Of Birth:")
        Nationality = input("Enter Your Nationality:")
        Psp_loc = Nationality
        Psp_No = input("Enter Your Passport Number:")
        Psp_Exp = input("Enter expiry Date Of Passport:")
        Phone_No = input("Enter Your Phone Number:")
        E_mail = input("Enter Your Email Address:")
        home_add = input("Enter Your Home address:")
        print('')
        print('')
        print('')
        print('PNR_NO :',2679054)
        print('Flight_NO :',747974298)
        Name=str(Fst_Name)+str(Sur_Name)
        print(Name)
        print('Rs',payment)
        print('Passport Location:',Psp_loc)
        print('Departure Date :',Depart_Date2)
        print('Return Date :',Return_Date2)
    elif opt == opt_2:

